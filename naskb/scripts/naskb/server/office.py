"""Office 在线预览（V2 零依赖简版，用户拍板）。

- docx：复用 batch._docx_flow_items 流式还原段落/表格/图片 → HTML；
  图片从 docx 压缩包 media 提取内嵌 base64（≤8 张），无外部依赖；
- xlsx：openpyxl 只读模式下前若干工作表 → HTML 表格（行列上限裁剪）；
- pptx：不在 V2 范围，返回 None（前端提示下载）。

产物直接内嵌 HTML 字符串返回，前端用 iframe srcdoc 沙箱显示。
"""
from __future__ import annotations

import base64
import html as _html
import os
import zipfile
from typing import Optional

_IMG_MIME = {"png": "image/png", "jpg": "image/jpeg", "jpeg": "image/jpeg",
             "gif": "image/gif", "bmp": "image/bmp", "webp": "image/webp",
             "svg": "image/svg+xml", "tif": "image/tiff", "tiff": "image/tiff"}
_MAX_DOCX_MB = 30
_MAX_IMAGES = 8
_SHEETS = 3
_XLSX_ROWS = 300
_XLSX_COLS = 40


def _esc(s: str) -> str:
    return _html.escape(s or "")


def _docx_media_uri(local: str, media: str) -> Optional[str]:
    """word/media 二进制作 data URI（找不到返回 None）。"""
    try:
        with zipfile.ZipFile(local) as z:
            path = "word/" + media.lstrip("/")
            data = z.read(path)
    except Exception:
        return None
    ext = media.rsplit(".", 1)[-1].lower() if "." in media else ""
    mime = _IMG_MIME.get(ext, "image/png")
    return f"data:{mime};base64,{base64.b64encode(data).decode('ascii')}"


def docx_to_html(local: str) -> Optional[str]:
    """docx → 简化 HTML（段落/标题启发 + 表格 + 少量内嵌图）。"""
    if os.path.getsize(local) > _MAX_DOCX_MB * 1024 * 1024:
        return None
    try:
        from ..common.batch import _docx_flow_items
        items = _docx_flow_items(local)
    except Exception:
        return None
    if not items:
        return None
    parts: list[str] = []
    img_count = 0
    for it in items:
        kind = it.get("kind")
        if kind == "text":
            txt = _esc(str(it["content"]).strip())
            if not txt:
                continue
            parts.append(f"<p>{txt}</p>")
        elif kind == "table":
            rows = [r for r in str(it["content"]).split("\n") if r]
            cells = ["<tr><td>{}</td></tr>".format(
                "</td><td>".join(_esc(c) for c in r.split(" | ")))
                for r in rows]
            parts.append("<table border='1' cellpadding='4' "
                         "style='border-collapse:collapse'>{}</table>"
                         .format("".join(cells)))
        elif kind == "image" and img_count < _MAX_IMAGES:
            uri = _docx_media_uri(local, str(it.get("media") or ""))
            if uri:
                parts.append("<p><img src='{}' style='max-width:100%'>"
                             "</p>".format(uri))
                img_count += 1
    return "<div>{}</div>".format("".join(parts))


def xlsx_to_html(local: str) -> Optional[str]:
    """xlsx → 前几工作表 HTML 表格（行列裁剪）。"""
    if os.path.getsize(local) > _MAX_DOCX_MB * 1024 * 1024:
        return None
    try:
        from openpyxl import load_workbook
        wb = load_workbook(local, read_only=True, data_only=True)
    except Exception:
        return None
    try:
        blocks: list[str] = []
        for ws in wb.worksheets[:_SHEETS]:
            rows_html: list[str] = []
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i >= _XLSX_ROWS:
                    break
                vals = [(
                    _esc(str(v)) if v is not None else "")
                    for v in row[:_XLSX_COLS]]
                if any(vals):
                    rows_html.append("<tr><td>{}</td></tr>".format(
                        "</td><td>".join(vals)))
            if rows_html:
                blocks.append(
                    "<h4>{}</h4><table border='1' cellpadding='4' "
                    "style='border-collapse:collapse'>{}</table>"
                    .format(_esc(str(ws.title)), "".join(rows_html)))
        return "<div>{}</div>".format("".join(blocks)) if blocks else None
    finally:
        try:
            wb.close()
        except Exception:
            pass


def render(config, local: str, ext: str) -> Optional[str]:
    """按扩展名渲染为 HTML；不支持/失败返回 None。"""
    e = (ext or "").lower()
    if e == "docx":
        return docx_to_html(local)
    if e == "xlsx":
        return xlsx_to_html(local)
    return None
